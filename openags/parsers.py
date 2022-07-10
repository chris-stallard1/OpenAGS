import re
import os

import xylib
import numpy as np
from scipy.interpolate import interp1d
from openpyxl import Workbook

from openags.util import KnownPeak


class SpectrumParser:
    """Parser for all Spectrum file formats"""
    def __init__(self, fname):
        self.fname = fname
        if self.fname.split(".")[-1].lower() == "spe":
            with open(self.fname) as f:
                firstLine = f.readline()
                if firstLine[0] == "$":
                    self.speFile = True
                else:
                    self.speFile = False
        else:
            self.speFile = False
    def getValues(self):
        if self.speFile:
            #Reader specifically for SPE files from Genie
            with open(self.fname) as f:
                text = f.read()
            sections = dict(section.split('\n', 1) for section in text.split('$')[1:])
            livetime, realtime = sections["MEAS_TIM:"].strip().split()
            livetime = float(livetime)
            realtime = float(realtime)
            startind, endind = sections["DATA:"].split("\n")[0].split()
            data = sections["DATA:"].split("\n")[1:-1]
            data = [int(i) for i in data]
            intercept, slope = sections["ENER_FIT:"].strip().split()[:2]
            intercept = float(intercept)
            slope = float(slope)
            energies = [intercept + i*slope for i in range(int(startind),int(endind)+1)]
            cps = [total/livetime for total in data]
            return {"livetime":livetime, "realtime":realtime, "energies":np.array(energies), "cps":np.array(cps)}
        else:
            #Otherwise just use xylib
            d = xylib.load_file(self.fname, '')
            block = d.get_block(0)
            meta = block.meta
            metaDict = {meta.get_key(i) : meta.get(meta.get_key(i)) for i in range(meta.size())}
            ncol = block.get_column_count()
            nrow = block.get_point_count()
            data = [[block.get_column(i).get_value(j) for j in range(nrow)] for i in range(1, ncol+1)]
            livetime = metaDict["live time (s)"]
            realtime = metaDict["real time (s)"]
            return {"livetime":livetime, "realtime":realtime, "energies":np.array(data[0]), "cps":np.array(data[1])/float(livetime)}

class StandardsFileParser:
    """Parser for standards files with peak locations, sensitivities, etc."""
    def __init__(self, k0_fname: str, eff_fname: str, flux=1):
        self.k0_fname = k0_fname
        self.eff_fname = eff_fname
        self.flux = flux
        self.peaks = None
    def extract_peaks(self, delayed):
        if self.peaks is not None:
            return self.peaks
        with open(self.k0_fname) as f:
            k0_lines = f.readlines()
        k0_headings = re.sub(r'[^\x00-\x7F]+', '', k0_lines[0]).strip().split(",") #Eliminate non-ASCII characters
        k0_lines = [line.split(",") for line in k0_lines[1:]]
        try:
            peakIndex = k0_headings.index("Energy (keV)")
            isoIndex = k0_headings.index("Isotope")
        except:
            raise ValueError("Bad k0 File Format.") # Those 2 headings are necessary
        #set of Regexes to test the other headings with
        reMass = re.compile(r"Mass \((\w+)\)")
        rek0 = re.compile (r"k0\w*")
        reHalfLife = re.compile(r"[Hh]alf-[Ll]ife \((\w+)\)")
        reDecayConstant = re.compile(r"[Dd]ecay [Cc]onstant \(1/(\w+)\)")

        #indexes to headings for each thing
        halfLifeIndex=None
        decayConstantIndex = None
        k0Index = None
        massIndex = None

        #Booleans telling us whether half-life/decay constant is used and whether sensitivity/mass is used.
        decayConstant = None
        useSensitivity = None
        
        #Units for mass and decay time
        unit = None
        decayUnit = None
        for i,h in enumerate(k0_headings):
            if reMass.match(h) is not None:
                useSensitivity = False
                massIndex = i
                unit = reMass.match(h).group(1)
            elif rek0.match(h) is not None:
                useSensitivity = True
                k0Index = i
                unit = reSens.match(h).group(1)
            elif reDecayConstant.match(h):
                decayConstant = True
                decayUnit = reDecayConstant.match(h).group(1)
                decayConstantIndex = i
            elif reHalfLife.match(h):
                decayConstant = False
                decayUnit = reHalfLife.match(h).group(1)
                halfLifeIndex = i

        with open(self.eff_fname) as f:
            eff_lines = f.readlines()
        eff_headings = re.sub(r'[^\x00-\x7F]+', '', eff_lines[0]).strip().lower().split(",") #Eliminate non-ASCII characters
        eff_lines = [line.split(",") for line in k0_lines[1:]]
        try:
            effEnergyIndex = eff_headings.index("energy (kev)")
            effIndex = eff_headings.index("efficiency")
        except:
            raise ValueError("Bad Efficiency Table Format.") # Those 2 headings are necessary

        eff_fn = interp1d([float(l[effEnergyIndex]) for l in eff_lines], [float(l[effIndex]) for l in eff_lines])

        #Create the known peak lists
        if useSensitivity is None:
            self.peaks = [KnownPeak(l[isoIndex],float(l[peakIndex])) for l in k0_lines]
        elif useSensitivity:
            self.peaks = [KnownPeak(l[isoIndex],float(l[peakIndex]), sensitivity = float(l[k0Index]) * flux * eff_fn(float(l[energyIndex])) / 4.987, unit=unit) for l in k0_lines]
        else:
            self.peaks = [KnownPeak(l[isoIndex],float(l[peakIndex]), mass = float(l[massIndex]), unit=unit) for l in k0_lines]
        
        #Adding the NAA parameters if needed
        if delayed:
            if decayConstant is None:
                raise ValueError("Delayed analysis, must provide half-life or decay constant in sensitivity file")
            elif decayConstant:
                for i in range(len(self.peaks)):
                    self.peaks[i].set_NAA_params(decayConstant = k0_lines[i][decayConstantIndex], unit=decayUnit)
            else:
                for i in range(len(self.peaks)):
                    self.peaks[i].set_NAA_params(halfLife = k0_lines[i][halfLifeIndex], unit=decayUnit)
        return self.peaks

class CSVWriter:
    """Writes a CSV file with name k0_fname to the results/projectID directory.
    
    The first row of the file contains column headings specified in headings.
    All remaining data is specified in data.
    """
    def __init__(self, projectID, fname, headings, data):
        self.fname = "./results/" + projectID + "/" + fname
        self.headings = headings
        self.data = data
    def write(self):
        with open(self.fname, "w") as f:
            f.seek(0)
            f.write(",".join(self.headings)+"\n")
            for line in self.data:
                try:
                    ld = [str(e) for e in line[0]]
                except:
                    ld = [str(e) for e in line]
                f.write(','.join(ld)+"\n")

class ExcelWriter:
    def __init__(self, projectID, projectTitle, allFilenames, headings, data):
        """Writes results from a project into an Excel File."""
        self.fname = "./results/" + projectID + "/" + projectTitle.replace(" ","_").replace("\\","").replace("/","") + ".xlsx"
        self.allFilenames = allFilenames
        self.headings = headings
        self.data = data
    def write(self):
        wb = Workbook()
        
        #First sheet w/compiled data
        ws = wb.active
        ws.title = "All Files"
        ws["A1"] = "Filename"
        for i in range(len(self.headings[0][0])):
            _ = ws.cell(row=1, column=i+2, value=self.headings[0][0][i])
        rowCount = 1
        for i in range(len(self.allFilenames)):
            for j in range(len(self.data[i][0])): #its data[i][0] and not data[i] because this writer only cares about 1 evaluator, the Mass/Sens one used in the program.
                _ = ws.cell(row=rowCount+1, column=1, value=os.path.split(self.allFilenames[i])[1])
                for k in range(len(self.data[i][0][j])):
                    _ = ws.cell(row=rowCount+1, column=k+2, value=self.data[i][0][j][k])
                rowCount += 1

        #remaining sheets
        for i in range(len(self.allFilenames)):
            newWs = wb.create_sheet(os.path.split(self.allFilenames[i])[1][:31]) #sheet name can't be longer than 31 chars
            for l in range(len(self.headings[0][0])):
                _ = newWs.cell(row=1, column=l+1, value=self.headings[0][0][l])
            for j in range(len(self.data[i][0])): #its data[i][0] and not data[i] because this writer only cares about 1 evaluator, the Mass/Sens one used in the program.
                for k in range(len(self.data[i][0][j])):
                    _ = newWs.cell(row=j+2, column=k+1, value=self.data[i][0][j][k])
        
        wb.save(self.fname)



